from openpyxl import load_workbook

def get_subjects_from_xlsx(SUBJECTS_FILE_PATH):
    wb =  load_workbook(filename=SUBJECTS_FILE_PATH, read_only=True)
    subjects_by_year = {}
    for sheet in wb:
        year_of_study = sheet.title
        subjects = []
        for row in sheet.iter_rows(min_row=1):
            subject = row[0].value
            if subject:
                subjects.append(subject)
        subjects_by_year[year_of_study] = subjects
    wb.close()
    return subjects_by_year

def arrange_subjects(subjects_by_year):
    for year_of_study, subjects in subjects_by_year.items():
        subjects = sorted(set(subject.strip()[:100] for subject in subjects))
        subjects_by_year[year_of_study] = subjects
    return subjects_by_year

def get_subjects_by_year(SUBJECTS_FILE_PATH):
    return arrange_subjects(get_subjects_from_xlsx(SUBJECTS_FILE_PATH))

# def get_students_from_xlsx(STUDENTS_FILE_PATH):
#     wb = load_workbook(filename=STUDENTS_FILE_PATH, read_only=True)
#     sheet = wb.active
#     students = []
#     for row in sheet.iter_rows(min_row=2):
#         promotion = row[0].value
#         code = row[2].value
#         last_name = row[5].value
#         first_name = row[6].value
#         mail = row[7].value
#         student = {
#             'promotion': promotion,
#             'code': code,
#             'last_name': last_name,
#             'first_name': first_name,
#             'mail': mail
#         }
#         students.append(student)
#     # Debug
#     for student in students:
#         print(f"Promotion: {student['promotion']}, Code: {student['code']}, Last name: {student['last_name']}, First name: {student['first_name']}, Mail: {student['mail']}")
#     wb.close()
#     return students

# DB_FILE=os.getenv('DB_FILE_PATH')

# def get_db():
#     if os.path.exists(DB_FILE):
#         try:
#             with open(DB_FILE, 'r') as f:
#                 return json.load(f)
#         except:
#             print(f"The file {DB_FILE} is empty or corrupted. Resetting.")
#             return {}
#     else:
#         print(f"The file {DB_FILE} does not exist. Creating.")
#         with open(DB_FILE, 'w') as f:
#             json.dump({}, f)
#         return {}

# def save_db(db):
#     with open(DB_FILE, 'w') as f:
#         json.dump(db, f, indent=4)